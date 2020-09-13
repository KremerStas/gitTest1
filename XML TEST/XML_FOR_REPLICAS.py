import openpyxl

with open('dialogiesTEST.txt', encoding='utf-8') as f:
    content = f.read()
dialogues = []
for dialogue_text in content.split('\n\n'):
    replicas = dialogue_text.split('\n')
    if len(replicas) >= 2:
        replicas = replicas[:2]
        replicas = [replica[2:] for replica in replicas]
        replicas[0] = replicas[0].strip()
        if replicas[0]:
            dialogues.append(tuple(replicas))
list(set(dialogues))


wb = openpyxl.Workbook()
sheet = wb.active
for replica in dialogues:
    first_cell = sheet.cell(row=dialogues.index(replica) + 1, column=1)
    first_cell.value = replica[0]
    second_cell = sheet.cell(row=dialogues.index(replica) + 1, column=2)
    second_cell.value = replica[1]
wb.save('REPLICAS.xls')